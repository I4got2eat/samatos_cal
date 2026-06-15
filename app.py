import streamlit as st
import fitz
import base64
import json
import pandas as pd
import re
import sqlite3
import io
from concurrent.futures import ThreadPoolExecutor, as_completed
from anthropic import Anthropic
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.utils import get_column_letter

# ── Config ────────────────────────────────────────────────────────────────────
# API key lives in .streamlit/secrets.toml (ANTHROPIC_API_KEY = "sk-ant-...")
API_KEY = st.secrets["ANTHROPIC_API_KEY"]
DB_PATH = r"C:\Users\Mykolas\claude_sessions\samata.db"

# Standard Lithuanian sąmata overhead rates
OVERHEAD = {
    "KDU":  0.08,    # Additional labor charges
    "PMD":  0.03,    # Additional materials value
    "PMZ":  0.03,    # Additional machinery value
    "SOD":  0.0179,  # Social insurance
    "STI":  0.09,    # Site overhead
    "PRI":  0.209,   # Indirect costs (applied to labor)
    "PLN":  0.05,    # Profit
    "PVM":  0.21,    # VAT
}


# ── Database ──────────────────────────────────────────────────────────────────
@st.cache_data
def load_resources():
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query("SELECT * FROM resources", conn)
    conn.close()
    return df


# ── PDF extraction (unchanged) ────────────────────────────────────────────────
def get_pdf_page_count(pdf_bytes):
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    n = len(doc)
    doc.close()
    return n


def extract_specific_pages_as_images(pdf_bytes, start_page, end_page, dpi=300):
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    target_images = []
    for page_num in range(max(0, start_page - 1), min(len(doc), end_page)):
        page = doc.load_page(page_num)
        pix = page.get_pixmap(dpi=dpi)
        base64_image = base64.b64encode(pix.tobytes("jpeg")).decode("utf-8")
        target_images.append({"page_num": page_num + 1, "base64_image": base64_image})
    doc.close()
    return target_images


def _parse_json_lenient(raw):
    """Parse JSON, repairing truncated output by trimming to the last complete
    object and closing the open array/object brackets."""
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        pass
    last = raw.rfind("}")
    while last != -1:
        candidate = raw[: last + 1]
        for suffix in ("]}", "]", ""):
            try:
                return json.loads(candidate + suffix)
            except json.JSONDecodeError:
                continue
        last = raw.rfind("}", 0, last)
    raise ValueError("Nepavyko apdoroti AI atsakymo (JSON)")


def extract_table_data(base64_image):
    client = Anthropic(api_key=API_KEY)
    prompt = """
    Extract the table content from this image.
    1. Analyze the document to identify hierarchical headers (Main Categories and Sub Categories).
    2. Flatten the table structure into a list of objects.
    3. For every row, identify the 'Current_Main_Section' and 'Current_Sub_Section' that applies to that data.
    4. Include all available columns in the row (e.g., Eil_Nr, Pavadinimas, Vnt, Kiekis, TS_skyrius, Pastabos).
    5. Return ONLY a JSON object with a single key 'table_data' containing the array of objects.

    Rules:
    - If a row is a section header, do not include it as a data row; use it to update the context for subsequent rows.
    - If a value is missing or spanned, infer it from the context of the identified section.
    - Do not use markdown, do not add conversational text.
    """
    # Streamed: the SDK requires streaming for requests with large max_tokens
    with client.messages.stream(
        model="claude-sonnet-4-6",
        max_tokens=32000,
        temperature=0,
        messages=[{
            "role": "user",
            "content": [
                {"type": "image", "source": {"type": "base64", "media_type": "image/jpeg", "data": base64_image}},
                {"type": "text", "text": prompt},
            ],
        }],
    ) as stream:
        raw = stream.get_final_text().strip()
    match = re.search(r"\{.*\}", raw, re.DOTALL)
    return _parse_json_lenient(match.group(0) if match else raw)


# ── AI estimator planning ─────────────────────────────────────────────────────
def plan_work_items(records):
    """Transform raw žiniaraštis rows into a professional estimator work list —
    the step a human sąmatininkas performs before pricing: normative naming,
    unit conversion, splitting complex items, adding implied works."""
    client = Anthropic(api_key=API_KEY)
    # Known norm catalog — planner should prefer these codes so the matcher can
    # resolve them deterministically from stored recipes.
    conn = sqlite3.connect(DB_PATH)
    norm_rows = conn.execute(
        "SELECT code, name, unit FROM work_norms ORDER BY code").fetchall()
    conn.close()
    norms_text = "\n".join(f"{c} | {n} | {u}" for c, n, u in norm_rows)
    prompt = f"""You are an experienced Lithuanian construction cost estimator (sąmatininkas) preparing
a work-item list for an Astera sąmata from a project "sąnaudų žiniaraštis".

Raw žiniaraštis rows extracted from the project PDF:
{json.dumps(records, ensure_ascii=False, indent=2)}

KNOWN NORM CATALOG (kodas | pavadinimas | matavimo vnt) — when a žiniaraštis work matches
one of these norms, you MUST use this exact code, name style and unit:
{norms_text}

Transform these rows into the final sąmata work-item list EXACTLY the way a professional
estimator does. Apply ALL of these rules:

1. RENAME each work into professional normative style (noun + verb form), e.g.
   "Ardomos perdangos plokštės" → "Perdangos plokščių demontavimas",
   "Grunto kasimas kanalams" → "Mechanizuotas grunto kasimas, pakraunant ir išvežant gruntą".
   Keep essential technical details (storis, betono klasė, profiliai) in the name.

2. CONVERT units and quantities to Lithuanian normative (normatyvų) units:
   - kg → t (kiekis ÷ 1000) for all steel/metal works
   - large areas in demolition, tankinimas, hidroizoliacija, dangos, plokščių montavimas:
     m2 → "100 m2" (kiekis ÷ 100) when the typical norm uses 100 m2
   - bulk earthworks: m3 → "100 m3" (kiekis ÷ 100)
   - long joints/siūlės: m → "100 m" (kiekis ÷ 100)
   - dual units like "m3/m2" or "m3/vnt" with Kiekis "a/b": the numbers pair
     POSITIONALLY — first number belongs to first unit, second to second
     (m3/vnt 48/24 means m3=48 OR vnt=24 — NEVER vnt=48). Choose the unit the
     norm is measured in and take ITS paired number.
   - "kiekis" MUST be expressed in the chosen "vnt". If you pick a norm whose unit
     is "100 m2" for a work given in m3, convert via the slab thickness
     (m3 ÷ storis_m ÷ 100); never leave the m3 number under a 100 m2 unit.
   - small quantities (mūras a few m3, betonavimas, vnt.) stay in their natural unit
   - perdangos/monolitinių konstrukcijų betonavimas keeps m3 norms (e.g. N6-109-1);
     N11P-15xx grindų norms (100 m2) are ONLY for grindys ant grunto
   - metal norms with unit "t" (N9-172 and similar) apply ONLY to metal works;
     mūro/betono griovimas uses mūro ardymo norms in m3 — never convert m3 of
     masonry into "t" via a metal norm

3. SPLIT complex ("kompleksiškai" priced) rows into separate work items:
   - "antikorozinis padengimas" / "dažymas C3(H)" of steel → two items:
     "Metalinių konstrukcijų gruntavimas" (t) + "Metalinių konstrukcijų dažymas" (t),
     tonnage taken from the related steel works
   - armavimas X kg/m3 in the description → a separate "armavimas tinklais" item
     (t, kiekis = m3 × X kg/m3 ÷ 1000) ONLY when the chosen betonavimo norm does
     NOT already include armatūrą. The norms F6-2-1, F6-2-5, N6-109-1 and F5-2
     ALREADY include armatūrą — never add a separate armavimas item for them.
     Grindys ant grunto (N11P-1502-1) does NOT include it — there the separate
     armavimas item (N11P-1508-2) is required.
   - the SAME work type appearing in several žiniaraštis rows must get the SAME
     norm code (e.g. visi kanalų dugno betonavimai → F6-2-1, visos sienutės →
     F6-2-5, regardless of which kanalas they belong to)
   - floors with "paviršius šlifuojamas padengiamas kietikliu" → add separate item
     "Betoninių dangų padengimas kietais užpildais" (100 m2, plotas iš storio: m3÷storis)
   - sąramų įrengimas with listed elements → a "Gelžbetoninių sąramų montavimas"
     work item (N7P-0306-3, vnt., total count) PLUS one material item per sąramos
     type with its count (e.g. Sąramos#SR16-37 → G-8218, Sąramos#SL16 → G-8245)

4. ADD implied works the žiniaraštis omits but every estimate includes:
   - ONE single "Statybinių šiukšlių išvežimas 10 km atstumu automobiliais-savivarčiais"
     (kodas R23-65, t) and ONE single "Sąvartyno mokestis" (kodas DDDD, t) — placed right
     after ALL demolition works, with tonnage summed over every demolition item.
     Estimate tonnage: concrete/gelžbetonis 2.5 t/m3, mūras 1.8 t/m3, metal from kg.
     Round the total to tens. Do NOT create separate transport/disposal items per work.

5. SUGGEST the most likely Lithuanian normative code for each work ("kodas"):
   F/N/R-style codes (e.g. F7-2-5, N9P-0101-1, N46-180-1, R23-65). This is orientacinis —
   choose the most plausible known code family for that work type. Typical picks:
   gelžbetoninių rygelių/plokščių demontavimas → F7-2-3 / F7-2-5 (m3);
   mechanizuotas grunto kasimas su išvežimu → F1-1-2 (100 m3);
   smėlio užpylimas aplink kanalus → N6-43 (m3);
   deformacinis profilis (TJS ir pan.) grindyse → N6P-0901-1 (100 m).

6. Quantities you do NOT convert must stay EXACTLY as in the source. Never drop a
   žiniaraštis row — every source row must map to at least one output item.

Return ONLY a JSON array (no markdown), each element:
{{
  "zin_nr": "source žiniaraštis row number(s), e.g. '1.1' or '2.4' (for added items: 'pridėta')",
  "kodas": "suggested normative code",
  "pavadinimas": "professional Lithuanian work name",
  "vnt": "normative unit (t, 100 m2, m3, vnt., ...)",
  "kiekis": <number in the normative unit>
}}"""
    with client.messages.stream(
        model="claude-sonnet-4-6",
        max_tokens=24000,
        temperature=0,
        messages=[{"role": "user", "content": prompt}],
    ) as stream:
        raw = stream.get_final_text().strip()
    m = re.search(r"\[.*\]", raw, re.DOTALL)
    return _parse_json_lenient(m.group(0) if m else raw)


# ── Deterministic post-planning validation ───────────────────────────────────
def _parse_num_lt(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


# Astera placeholder codes shared by unrelated custom items — never treat as norms
DUMMY_NORM_CODES = {"DDDD"}


def refine_planned_items(planned, source_rows):
    """Cross-check AI-planned quantities/units against the source žiniaraštis rows.
    Fixes the classic failure modes deterministically: mixed dual-unit picks
    (m3/vnt 48/24), missed scaling (m3 value under a '100 m2' unit), kg→t,
    forces norm units when the suggested code exists in work_norms, and drops
    armavimas split-items whose betonavimo norm already includes armatūra."""
    conn = sqlite3.connect(DB_PATH)
    norm_units = dict(conn.execute("SELECT code, unit FROM work_norms").fetchall())
    arm_norms = {r[0] for r in conn.execute(
        "SELECT DISTINCT norm_code FROM norm_resources WHERE resource_name LIKE '%rmat%'")}
    conn.close()
    for c in DUMMY_NORM_CODES:
        norm_units.pop(c, None)

    # zin_nr -> source rows (duplicate Eil_Nr possible across sections)
    by_nr = {}
    for row in source_rows:
        nr = str(row.get("Eil_Nr") or row.get("eil_nr") or "").strip()
        by_nr.setdefault(nr, []).append(row)

    def src_parts(row):
        units = [u.strip() for u in str(row.get("Vnt") or row.get("vnt") or "").split("/") if u.strip()]
        qtys  = [_parse_num_lt(q) for q in str(row.get("Kiekis") or row.get("kiekis") or "").split("/")]
        text  = str(row.get("Pavadinimas") or row.get("pavadinimas") or "")
        return units, qtys, text

    def candidates(units, qtys, text, target_unit):
        tn = target_unit.replace(" ", "").lower().rstrip(".")
        out = []
        pairs = list(zip(units, qtys))
        for u, q in pairs:
            if q is None:
                continue
            un = u.replace(" ", "").lower().rstrip(".")
            if un == tn:
                out.append(q)
        for u, q in pairs:
            if q is None:
                continue
            un = u.replace(" ", "").lower().rstrip(".")
            if tn == "t" and un == "kg":
                out.append(q / 1000)
            if tn.startswith("100") and un == tn[3:]:
                out.append(q / 100)
        if tn == "100m2":  # m3 with known thickness -> area
            m3 = next((q for u, q in pairs if q is not None and u.replace(" ", "").lower() == "m3"), None)
            mt = re.search(r"t\s*[=\-]\s*(\d+)", text)
            if m3 and mt and int(mt.group(1)) >= 50:
                out.append(m3 / (int(mt.group(1)) / 1000) / 100)
        return out

    for it in planned:
        name_l = str(it.get("pavadinimas") or "").lower()
        # fixed conventions for the standard added items
        if "sąvartyno" in name_l:
            it["kodas"], it["vnt"] = "DDDD", "t"
            continue
        if "šiukšlių išvežimas" in name_l:
            it["kodas"], it["vnt"] = "R23-65", "t"
            continue

        code = str(it.get("kodas") or "").strip()
        if code in norm_units and norm_units[code]:
            it["vnt"] = norm_units[code]  # norm unit is authoritative

        nr = str(it.get("zin_nr") or "").strip()
        rows = by_nr.get(nr, [])
        if not rows or "armavim" in name_l:
            continue  # derived/added items have no direct source quantity
        qty = _parse_num_lt(it.get("kiekis"))
        if qty is None:
            continue
        best = None
        for row in rows:
            units, qtys, text = src_parts(row)
            cands = candidates(units, qtys, text, str(it.get("vnt") or ""))
            if any(c is not None and abs(c - qty) <= 0.02 * max(abs(c), 1e-9) for c in cands):
                best = qty  # planner value is consistent with a source candidate
                break
            if cands and best is None:
                best = cands[0]
        if best is not None and abs(best - qty) > 0.02 * max(abs(best), 1e-9):
            it["kiekis"] = round(best, 4)

    # Drop armavimas split-items double-counting steel: if the betonavimo item
    # from the same žiniaraštis row uses a norm whose recipe already includes
    # armatūra (F6-2-x, N6-109-1, F5-2 …), the separate armavimas row is wrong.
    parent_code = {}
    for it in planned:
        if "armavim" not in str(it.get("pavadinimas") or "").lower():
            parent_code.setdefault(str(it.get("zin_nr") or ""), str(it.get("kodas") or ""))
    result = []
    for it in planned:
        name_l = str(it.get("pavadinimas") or "").lower()
        if "armavim" in name_l and parent_code.get(str(it.get("zin_nr") or "")) in arm_norms:
            continue
        result.append(it)
    return result


# ── AI matching ───────────────────────────────────────────────────────────────
MATCH_BATCH_SIZE = 8  # items per API call to stay well within output token limit

def _match_batch(client, items_json, resources_text):
    prompt = f"""You are a strict Lithuanian construction cost estimator working with the Astera sąmata system.

Below are construction specification items extracted from a client project PDF:
{items_json}

Below is the COMPLETE price database you are allowed to use (format: code | name | unit | unit_price | category):
{resources_text}

STRICT RULES — read carefully:
1. You may ONLY use resources that exist verbatim in the database above. Do NOT invent codes, names, prices, or resources.
2. If a spec item cannot be matched to ANY resource in the database with reasonable confidence, set "status": "no_match" and leave "matched_resources" as an empty array.
3. A "reasonable match" means the resource clearly corresponds to the spec item material/labor/machinery. Do NOT force a match just to fill the field.
4. You may combine multiple database resources for one spec item ONLY if each individual resource genuinely applies.
5. Never hallucinate a unit_price — copy it exactly from the database row above.
6. Skip items where Kiekis is null, missing, or "-".
7. All costs in EUR.
8. For "pavadinimas": copy the FULL Pavadinimas text from the spec item EXACTLY as it appears — do NOT shorten, summarise, or truncate it in any way.
9. The spec unit ("vnt") may be a scaled normative unit such as "t", "100 m2", "100 m3" or "100 m".
   "qty_per_spec_unit" must then be the resource quantity per 1 SUCH unit (e.g. darbo valandos per 1 toną,
   per 100 m2 — typical Lithuanian normatyvai values), NOT per kg or per single m2.

Return a JSON array where each element has this exact structure:
{{
  "eil_nr": "item number from spec",
  "pavadinimas": "FULL Pavadinimas text copied verbatim — every word, do not shorten",
  "vnt": "unit of measure",
  "kiekis": <total quantity as number>,
  "status": "matched" | "no_match",
  "matched_resources": [
    {{
      "code": "resource code copied exactly from the database",
      "name": "resource name copied exactly from the database",
      "unit": "resource unit copied exactly from the database",
      "unit_price": <price copied exactly from the database>,
      "qty_per_spec_unit": <resource quantity per 1 spec unit — your estimate>,
      "total_qty": <kiekis × qty_per_spec_unit>,
      "total_cost": <total_qty × unit_price>,
      "category": "labor|material|machinery"
    }}
  ],
  "total_labor": <sum of labor costs, 0 if no_match>,
  "total_materials": <sum of material costs, 0 if no_match>,
  "total_machinery": <sum of machinery costs, 0 if no_match>,
  "total_cost": <grand total, 0 if no_match>
}}

Return ONLY a valid JSON array, no markdown, no explanation.
"""
    with client.messages.stream(
        model="claude-sonnet-4-6",
        max_tokens=16000,
        temperature=0,
        messages=[{"role": "user", "content": prompt}],
    ) as stream:
        raw = stream.get_final_text().strip()
    m = re.search(r"\[.*\]", raw, re.DOTALL)
    return _parse_json_lenient(m.group(0) if m else raw)


def _match_one_via_norm(cur, item):
    """If the item's suggested norm code has a stored recipe (norm_resources),
    build the resource breakdown deterministically — no AI guessing.
    Returns the matched item dict, or None when no recipe exists."""
    code = str(item.get("eil_nr") or "").strip()
    kiekis = item.get("kiekis") or 0
    if not code or not kiekis or code in DUMMY_NORM_CODES:
        return None
    norm = cur.execute("SELECT name, unit FROM work_norms WHERE code = ?", (code,)).fetchone()
    if not norm:
        return None
    recipe = cur.execute(
        """SELECT nr.resource_code, nr.resource_name, nr.unit, nr.norm_per_unit,
                  nr.category, COALESCE(r.unit_price, 0)
           FROM norm_resources nr
           LEFT JOIN resources r ON r.code = nr.resource_code
           WHERE nr.norm_code = ?""", (code,)).fetchall()
    if not recipe:
        return None
    mres, totals = [], {"labor": 0.0, "material": 0.0, "machinery": 0.0}
    for rc, rn, ru, npu, cat, price in recipe:
        qty = (npu or 0) * kiekis
        cost = qty * (price or 0)
        mres.append({
            "code": rc, "name": rn, "unit": ru, "unit_price": price or 0,
            "qty_per_spec_unit": npu or 0, "total_qty": qty,
            "total_cost": round(cost, 2), "category": cat,
        })
        if cat in totals:
            totals[cat] += cost
    return {
        **item,
        "vnt": norm[1] or item.get("vnt"),   # norm unit is authoritative
        "status": "matched",
        "matched_resources": mres,
        "total_labor":     round(totals["labor"], 2),
        "total_materials": round(totals["material"], 2),
        "total_machinery": round(totals["machinery"], 2),
        "total_cost":      round(sum(totals.values()), 2),
    }


def match_items_to_resources(parsed_df, resources_df):
    client = Anthropic(api_key=API_KEY)

    resources_text = "\n".join(
        f"{r['code']} | {r['name']} | {r['unit']} | {r['unit_price'] if r['unit_price'] else 'N/A'} EUR | {r['category']}"
        for _, r in resources_df.iterrows()
    )

    all_records = parsed_df.to_dict(orient="records")

    # 1) Deterministic pass: items whose norm code has a stored recipe in the DB
    final = [None] * len(all_records)
    pending = []
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    for idx, rec in enumerate(all_records):
        hit = _match_one_via_norm(cur, rec)
        if hit is not None:
            final[idx] = hit
        else:
            pending.append((idx, rec))
    conn.close()

    if not pending:
        return final

    # 2) AI pass for the remaining items
    records = [rec for _, rec in pending]
    batches = [
        (i, records[i : i + MATCH_BATCH_SIZE])
        for i in range(0, len(records), MATCH_BATCH_SIZE)
    ]

    # Run all batches in parallel — total time ≈ one API call regardless of page count
    batch_results = {}
    with ThreadPoolExecutor(max_workers=min(len(batches), 10)) as pool:
        futures = {
            pool.submit(_match_batch, client, json.dumps(batch, ensure_ascii=False, indent=2), resources_text): idx
            for idx, batch in batches
        }
        for future in as_completed(futures):
            idx = futures[future]
            batch_results[idx] = future.result()

    # Reassemble AI results in original order, then merge into final positions
    ai_results = []
    for idx, _ in batches:
        ai_results.extend(batch_results[idx])
    for (pos, rec), item in zip(pending, ai_results):
        final[pos] = item
    # Any leftover slots (AI returned fewer items) become explicit no-matches
    for pos, rec in pending:
        if final[pos] is None:
            final[pos] = {**rec, "status": "no_match", "matched_resources": [],
                          "total_labor": 0, "total_materials": 0,
                          "total_machinery": 0, "total_cost": 0}
    return [x for x in final if x is not None]


# ── Cost rollup ───────────────────────────────────────────────────────────────
def compute_totals(matched_items):
    labor     = sum(i.get("total_labor", 0) or 0 for i in matched_items)
    materials = sum(i.get("total_materials", 0) or 0 for i in matched_items)
    machinery = sum(i.get("total_machinery", 0) or 0 for i in matched_items)
    direct    = labor + materials + machinery

    kdu  = labor * OVERHEAD["KDU"]
    pmd  = materials * OVERHEAD["PMD"]
    pmz  = machinery * OVERHEAD["PMZ"]
    sod  = (labor + kdu) * OVERHEAD["SOD"]
    t2   = direct + kdu + pmd + pmz + sod

    sti  = t2 * OVERHEAD["STI"]
    t3   = t2 + sti

    pri  = labor * OVERHEAD["PRI"]
    pln  = t3 * OVERHEAD["PLN"]
    t4   = t3 + pri + pln

    pvm  = t4 * OVERHEAD["PVM"]
    t5   = t4 + pvm

    return {
        "labor": labor, "materials": materials, "machinery": machinery,
        "direct": direct,
        "KDU": kdu, "PMD": pmd, "PMZ": pmz, "SOD": sod, "t2": t2,
        "STI": sti, "t3": t3,
        "PRI": pri, "PLN": pln, "t4": t4,
        "PVM": pvm, "t5": t5,
    }


# ── Excel generation ──────────────────────────────────────────────────────────

# Column layout — matches completed_estimate_sheet.xlsx exactly
# A=Nr, B=Name, C=Code, D=Unit, E=Norma, F=Koef(narrow),
# G=Kaina, H=Kiekis, I=Suma, J=Darbas, K=Medžiagos, L=Mechanizmai, M=Subrangovai
COL_WIDTHS = {
    "A": 4.7109375,  "B": 30.7109375, "C": 9.7109375,  "D": 8.0,        "E": 7.0,
    "F": 1.42578125,
    "G": 10.85546875,"H": 8.42578125, "I": 15.140625,
    "J": 11.42578125,"K": 9.7109375,  "L": 11.85546875,
    # hidden helper columns: M=Subrangovai, N=St koef, O=coeffs, P=St kodas, Q=markers
    "M": 30.7109375, "N": 8.0, "O": 8.0,
    "P": 10.7109375, "Q": 30.7109375,
    # R/S/T/U — visible markup columns
    "R": 15.28515625, "T": 17.28515625, "U": 16.5703125,
    "V": 9.140625, "Z": 9.140625, "AA": 9.140625,
    "AQ": 57.7109375, "AR": 9.140625, "BC": 9.140625,
}
# Columns hidden in the reference Astera file (raw XML: F; M–O range; P; Q; V–AR span).
# Visible columns are exactly A–E, G–L, R, S, T, U.
HIDDEN_COLS = (
    ["F", "M", "N", "O", "P", "Q"]
    + [get_column_letter(i) for i in range(22, 56)]  # V (22) … BC (55) helper span
)

def lt_num(value):
    """Format number as Lithuanian amount string: 8 440,24"""
    if value is None or value == "":
        return ""
    n = float(value)
    # Format with 2 decimals, space as thousands sep, comma as decimal
    formatted = f"{abs(n):,.2f}".replace(",", "X").replace(".", ",").replace("X", " ")
    return f"-{formatted}" if n < 0 else formatted


RED_FILL      = PatternFill("solid", fgColor="FF9999")
# Astera template fills (8-hex ARGB to match the reference exactly)
TEAL_FILL     = PatternFill("solid", fgColor="FF00C0C0")  # section header / total / row 11
LAVENDER_FILL = PatternFill("solid", fgColor="FFC0C0FF")  # work-item rows (A–M)
GREEN_FILL    = PatternFill("solid", fgColor="FF00C000")  # editable Kiekis cell
ITEM_TOP      = Border(top=Side(style="medium"))
ITEM_TOP_RED  = Border(top=Side(style="medium", color="CC0000"))
HDR_BORDER    = Border(top=Side(style="medium"), left=Side(style="medium"), right=Side(style="medium"))
DOUBLE_BOX    = Border(top=Side(style="double"), bottom=Side(style="double"),
                       left=Side(style="double"), right=Side(style="double"))
HDR_BORDER_FULL = Border(top=Side(style="medium"), bottom=Side(style="medium"),
                          left=Side(style="medium"), right=Side(style="medium"))

def _font(bold=False, size=10):
    return Font(name="Times New Roman", bold=bold, size=size)

def _align(h="left", v="top", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _calc_overhead_coeffs():
    """
    Compute the cumulative overhead multiplier per cost category (excl. VAT).
    Each coefficient answers: "for €1 of direct cost in this category, what is
    the €-value after all overheads are applied (excl. VAT)?"
    These are stored in O/P/Q of the grand-total #1 row and referenced by the
    R/S/T formulas for each work item.
    """
    KDU = OVERHEAD["KDU"];  SOD = OVERHEAD["SOD"];  STI = OVERHEAD["STI"]
    PLN = OVERHEAD["PLN"];  PRI = OVERHEAD["PRI"]
    PMD = OVERHEAD["PMD"];  PMZ = OVERHEAD["PMZ"]
    # Labor — Astera chain: PRI on (labor+KDU) joins after indexation, PLN on top of both:
    #   J#4 = (1+PLN) * ( (1+KDU)*PRI + (1+KDU+(1+KDU)*SOD)*(1+STI) ) * J#1
    coeff_L  = round((1 + PLN) * ((1 + KDU) * PRI + (1 + KDU + (1 + KDU) * SOD) * (1 + STI)), 6)
    # Materials: PMD on materials, STI on t2, PLN on t3
    coeff_M  = round((1 + PMD) * (1 + STI) * (1 + PLN), 6)
    # Machinery: same structure as materials (PMZ replaces PMD, same rates)
    coeff_Ma = round((1 + PMZ) * (1 + STI) * (1 + PLN), 6)
    return coeff_L, coeff_M, coeff_Ma

_COEFF_L, _COEFF_M, _COEFF_MA = _calc_overhead_coeffs()


def generate_samata_excel(matched_items, project_name=""):
    from datetime import date as _date
    import uuid

    wb = Workbook()
    ws = wb.active
    ws.title = "Sąmata"

    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width
    for col_letter in HIDDEN_COLS:
        ws.column_dimensions[col_letter].hidden = True

    # ── Pre-compute grand-total row numbers so rows 7/8 can reference them ────
    # Layout: row 12 = Skyrius, row 13+ = item blocks, section total,
    # then the 15-row overhead block (#1 … #5) immediately after — no blank
    # row in between, exactly like the Astera file (679 → 680).
    _data_rows = 1  # Skyrius header
    for _item in matched_items:
        _data_rows += 1 + len(_item.get("matched_resources", []))
    _data_rows += 1  # section total
    _r1 = 12 + _data_rows        # "Iš viso #1" row
    _r5 = _r1 + 14               # "Iš viso #5" row  (14 rows after #1)

    EUR_FMT = '#,##0.00\\ [$€-1];\\-#,##0.00\\ [$€-1]'

    def s(row, col, value, bold=False, size=9.75, h="left", v="top", wrap=True, red=False, fmt=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = _font(bold=bold, size=size)
        c.alignment = _align(h=h, v=v, wrap=wrap)
        if red:
            c.fill = RED_FILL
        if fmt:
            c.number_format = fmt
        return c

    # ── Program-helper rows 1–5 (same as the Astera file) ─────────────────────
    # I1 is the labor wage coefficient referenced by sub-item F formulas (=$I$1).
    ws.row_dimensions[1].height = 16.9
    _n_items       = len(matched_items)
    _last_data_row = _r1 - 2  # last sub-item row
    for col, val in [(1,0),(2,0),(3,_last_data_row),(4,"Eil.kiekis"),(5,_n_items),
                     (6,"Darbų kiekis"),(7,1),(8,1),(9,1),(10,2.5),(11,1),(12,False),(13,4)]:
        s(1, col, val)
    for col, val in [(1,"A"),(7,1),(8,1),(9,1),(10,"i8"),(12,"i1")]:
        s(2, col, val)
    s(3, 7, 1)
    s(4, 10, "Val. darbo užmokestis")
    s(4, 12, 6)
    s(5, 10, "Lito kursas")
    s(5, 11, 1)
    s(5, 12, "0,2896")

    # ── Header rows 6–8 ────────────────────────────────────────────────────────
    ws.row_dimensions[6].height = 25.5
    s(6, 2, project_name or "Projektas", bold=True, wrap=True)
    s(6, 4, "L o k a l i n ė   s ą m a t a   N r.", bold=True, h="left")

    ws.row_dimensions[7].height = 26.25
    s(7, 2, "", bold=True, wrap=True)
    s(7, 4, f"Sudaryta {_date.today().strftime('%Y.%m')} kainų lygiu.", bold=False, h="left")
    # I7: consistency check — sub-item sum minus grand total #1 (must equal 0)
    s(7, 9, f"=ROUND(SUM(I10:OFFSET(I{_r1},-1,0))-I{_r1},2)")

    ws.row_dimensions[8].height = 15.0
    s(8, 2, project_name or "", bold=True)
    s(8, 7, "Iš viso už", bold=True)
    # I8 — references grand total #5 via OFFSET exactly like the Astera file
    c_tot           = ws.cell(row=8, column=9, value=f"=OFFSET($I${_r5 + 1},-1,0)")
    c_tot.font      = _font(bold=True)
    c_tot.alignment = _align(h="center", v="center")
    c_tot.border    = DOUBLE_BOX
    c_tot.number_format = EUR_FMT

    ws.row_dimensions[10].height = 27.75
    for col, label in [
        (1,"Nr."),(2,"Darbo pavadinimas"),(3,"Kodas"),(4,"Mat. vnt"),
        (5,"Norma"),(6,"Koef"),(7,"Kaina"),(8,"Kiekis"),(9,"Suma"),
        (10,"Darbas"),(11,"Medžiagos"),(12,"Mechanizmai"),(13,"Subrangovai"),
    ]:
        c           = ws.cell(row=10, column=col, value=label)
        c.font      = _font(bold=True)
        c.alignment = _align(h="center", v="center")
        c.border    = HDR_BORDER
    # Hidden helper-column headers (no borders, same as the real file)
    s(10, 14, "St koef",  bold=True, h="center", v="center")
    s(10, 16, "St kodas", bold=True, h="center", v="center")
    # R/S/T/U headers use all 4 borders (top+bottom+left+right), matching the real Astera file
    for col, label in [
        (18, "Vieneto kaina su prisk. be PVM"),
        (19, "Vieneto kaina su prisk. su PVM"),
        (20, "Suma su prisk. be PVM visam kiekiui"),
        (21, "Vieneto kaina be priskaitymų"),
    ]:
        c           = ws.cell(row=10, column=col, value=label)
        c.font      = _font(bold=True)
        c.alignment = _align(h="center", v="center")
        c.border    = HDR_BORDER_FULL

    ws.freeze_panes = "A11"
    ws.row_dimensions[11].height = 12.75
    # Row 11 — teal separator band beneath the column headers (A–M)
    for col in range(1, 14):
        ws.cell(row=11, column=col).fill = TEAL_FILL

    # ── Section header (row 12) ───────────────────────────────────────────────
    r = 12
    section_name       = "Darbai"
    section_header_row = r
    ws.row_dimensions[r].height = 12.75
    for col in range(1, 14):
        ws.cell(row=r, column=col).fill = TEAL_FILL
    s(r, 2, "Skyrius",    bold=True, v="top", wrap=False)
    s(r, 3, section_name, bold=True, v="top", wrap=False)
    s(r, 17, "SP")  # hidden Q-column section marker, as in the Astera file
    r += 1

    # ── Work items with live formulas ─────────────────────────────────────────
    item_no = 1

    for item in matched_items:
        no_match  = item.get("status") == "no_match" or not item.get("matched_resources")
        kiekis    = item.get("kiekis") or 0
        resources = item.get("matched_resources", [])
        n_subs    = len(resources)
        item_row  = r
        last_sub  = item_row + n_subs  # row of the last sub-item

        # Row height: auto-size from name length
        _nl = len(item.get("pavadinimas", ""))
        ws.row_dimensions[r].height = max(25.5, -(-_nl // 30) * 13.5)

        border = ITEM_TOP_RED if no_match else ITEM_TOP
        for col in range(1, 14):
            ws.cell(row=r, column=col).border = border
            if not no_match:
                ws.cell(row=r, column=col).fill = LAVENDER_FILL
        for col in [18, 19, 20, 21]:
            ws.cell(row=r, column=col).border = border
        if not no_match:
            ws.cell(row=r, column=8).fill = GREEN_FILL  # H = editable Kiekis

        # Columns A–D (always plain text/numbers)
        s(r,  1, item_no,                     bold=True, size=9.75, h="center", v="top", red=no_match)
        s(r,  2, item.get("pavadinimas", ""), bold=True, size=9.75, h="left",   v="top", red=no_match)
        s(r,  3, item.get("eil_nr", ""),      bold=True, size=9.75, h="left",   v="top", red=no_match)
        s(r,  4, item.get("vnt", ""),         bold=True, size=9.75, h="center", v="top", red=no_match)
        s(r,  5, None, bold=True, size=9.75, red=no_match)  # Norma — blank for work items
        s(r,  6, None, bold=True, size=9.75, red=no_match)  # Koef  — blank

        # Column H: quantity (editable number)
        s(r,  8, kiekis, bold=True, size=9.75, h="right", v="top", red=no_match)

        if no_match:
            s(r,  7, None, bold=True, size=9.75, h="right", v="top", red=True)
            s(r,  9, "NA", bold=True, size=9.75, h="right", v="top", red=True)
            s(r, 10, "NA", bold=True, size=9.75, h="right", v="top", red=True)
            s(r, 11, "NA", bold=True, size=9.75, h="right", v="top", red=True)
            s(r, 12, "NA", bold=True, size=9.75, h="right", v="top", red=True)
            s(r, 13, None, bold=True, size=9.75, red=True)
            r += 1
        else:
            if n_subs:
                # G: unit price — CSE array formula Σ(norma × kaina) over the sub-rows
                gc = ws.cell(row=r, column=7)
                gc.value = ArrayFormula(f"G{r}",
                    f"=SUM(OFFSET(E{r},1,0):E{last_sub}*OFFSET(G{r},1,0):G{last_sub})")
                gc.font          = _font(bold=True, size=9.75)
                gc.alignment     = _align(h="right", v="top")
                gc.number_format = "0.00"
                # I/J/K/L/M: TEXT() formulas render as strings, so the column-wide
                # SUMs in section/grand totals skip them (no double counting)
                s(r,  9, f'=TEXT(VALUE(J{r}+K{r}+L{r}+M{r}),"0,00")',
                  bold=True, size=9.75, h="right", v="top", fmt="#,##0.00")
                for cl, col in [("J", 10), ("K", 11), ("L", 12), ("M", 13)]:
                    s(r, col, f'=TEXT(SUM(OFFSET({cl}{r},1,0):{cl}{last_sub}),"0,00")',
                      bold=True, size=9.75, h="right", v="top")
            else:
                for col in [7, 9, 10, 11, 12, 13]:
                    s(r, col, None, bold=True, size=9.75, h="right", v="top")
            # N: sub-row count; P: blank marker (hidden helper columns)
            s(r, 14, n_subs, bold=True, size=9.75)
            s(r, 16, "",     bold=True, size=9.75)
            # R/S/T/U — bottom-up: work item sums its sub-rows (Astera template).
            # O${_r1}/P/Q = category overhead coeffs; O${r_kdu}=1.21 VAT factor.
            _r_kdu = _r1 + 1
            if n_subs:
                first_sub = r + 1
                s(r, 18, f"=ROUND(SUM(R{first_sub}:R{last_sub}),2)",
                  bold=True, size=9.75, h="center", v="top")
                s(r, 19, f"=ROUND(SUM(O${_r_kdu}*R{r}),2)",
                  bold=True, size=9.75, h="center", v="top")
                s(r, 20, f"=ROUND(R{r}*H{r},2)",
                  bold=True, size=9.75, h="center", v="top")
                s(r, 21, f"=ROUND(SUM(U{first_sub}:U{last_sub}),2)",
                  bold=True, size=9.75, h="center", v="top")
            # AQ: per-item GUID (hidden), as emitted by Astera
            s(r, 43, "{" + str(uuid.uuid4()).upper() + "}")
            r += 1

            # ── Sub-item rows ─────────────────────────────────────────────────
            for res in resources:
                sr  = r   # sub-item row number
                cat = res.get("category", "")
                s(r,  1, None,                         bold=False, size=9.75, h="center", v="top")
                s(r,  2, res.get("name", ""),          bold=False, size=9.75, h="left",   v="top")
                s(r,  3, res.get("code", ""),          bold=False, size=9.75, h="left",   v="top")
                s(r,  4, res.get("unit", ""),          bold=False, size=9.75, h="center", v="top")
                s(r,  5, res.get("qty_per_spec_unit"), bold=False, size=9.75, h="right",  v="top")  # E Norma
                # F/G: labor price = base rate × wage coefficient ($I$1);
                # materials/machinery carry a plain number, like the Astera file
                if cat == "labor":
                    s(r, 6, "=$I$1", bold=False, size=9.75, h="right", v="top")
                    s(r, 7, f"={res.get('unit_price') or 0}*F{sr}",
                      bold=False, size=9.75, h="right", v="top")
                else:
                    s(r, 7, res.get("unit_price"), bold=False, size=9.75, h="right", v="top")
                # H: quantity = norm × parent kiekis
                s(r,  8, f"=E{sr}*H{item_row}",       bold=False, size=9.75, h="right",  v="top")
                # I: total cost = unit_price × quantity
                s(r,  9, f"=ROUND(G{sr}*H{sr},2)",    bold=False, size=9.75, h="right",  v="top")
                # J/K/L: route cost to correct category column
                s(r, 10, f"=I{sr}" if cat == "labor"    else None, bold=False, size=9.75, h="right", v="top")
                s(r, 11, f"=I{sr}" if cat == "material" else None, bold=False, size=9.75, h="right", v="top")
                s(r, 12, f"=I{sr}" if cat == "machinery"else None, bold=False, size=9.75, h="right", v="top")
                # N/O/P hidden helpers: St koef, category code (1/20/70), St kodas
                s(r, 14, 1, bold=False, size=9.75)
                s(r, 15, {"labor": 1, "material": 20, "machinery": 70}.get(cat, 20),
                  bold=False, size=9.75, h="center")
                s(r, 16, str(res.get("code", "")), bold=False, size=9.75, h="center")
                # R/S/T/U sub-row formulas (Astera template). Coefficient column by
                # category: labor→O, material→P, machinery→Q (all at #1 row _r1).
                _coef = {"labor": "O", "material": "P", "machinery": "Q"}.get(cat, "P")
                _r_kdu = _r1 + 1
                s(r, 18, f"=IF(H{item_row}<>0,G{sr}*H{sr}*{_coef}${_r1}/H{item_row},0)",
                  bold=False, size=9.75, h="right", v="top")
                s(r, 19, f'=TEXT(ROUND(SUM(O${_r_kdu}*R{sr}),2),"0,00")',
                  bold=False, size=9.75, h="right", v="top")
                s(r, 20, f'=TEXT(ROUND(ROUND(R{sr},2)*H{item_row},2),"0,00")',
                  bold=False, size=9.75, h="right", v="top")
                s(r, 21, f"=IF(H{item_row}<>0,G{sr}*H{sr}/H{item_row},0)",
                  bold=False, size=9.75, h="right", v="top")
                r += 1

        item_no += 1

    # ── Section total ─────────────────────────────────────────────────────────
    section_total_row = r
    for col in range(1, 14):
        ws.cell(row=r, column=col).fill = TEAL_FILL
    s(r, 2, "Iš viso už skyrių", bold=True)
    s(r, 3, f"=C{section_header_row}", bold=True)
    # TEXT() range sums — work-item rows are strings, so only sub-item numbers count
    for cl, col in [("I", 9), ("J", 10), ("K", 11), ("L", 12), ("M", 13)]:
        s(r, col, f'=TEXT(SUM({cl}{section_header_row}:OFFSET({cl}{r},-1,0)),"0,00")',
          bold=True, h="right")
    s(r, 17, "US")  # hidden Q-column section-total marker
    # T: sums work-item T values only (sub-items have TEXT strings, skipped by SUM)
    s(r, 20, f'=TEXT(SUM(T{section_header_row}:T{r - 1}),"0,00")',
      bold=False, h="center")
    r += 1  # overhead block follows immediately — no blank row, as in the file

    # ── Grand totals — identical structure to Astera rows 680–694 ─────────────
    # Labels sit in column C, editable rates in J/K/L of the same row, hidden
    # code markers in column P, and the I column carries the € amounts.
    # #1 ── direct costs
    r1 = r
    assert r1 == _r1, f"#1 row mismatch: expected {_r1}, got {r1}"
    s(r, 2, "Iš viso #1", bold=True)
    s(r,  9, f"=ROUND(J{r1}+K{r1}+L{r1},2)", bold=True, fmt=EUR_FMT)
    # Column-wide SUMs from row 10 — TEXT cells in work-item/section rows are skipped
    s(r, 10, f"=SUM(J$10:OFFSET(J{r1},-1,0))", fmt="0.00")
    s(r, 11, f"=SUM(K$10:OFFSET(K{r1},-1,0))", fmt="0.00")
    s(r, 12, f"=SUM(L$10:OFFSET(L{r1},-1,0))", fmt="0.00")
    s(r, 14, 8)
    # O/P/Q: overhead coefficients referenced by the work-item R/S/T formulas
    s(r, 15, _COEFF_L)
    s(r, 16, _COEFF_M)
    s(r, 17, _COEFF_MA)
    r += 1

    # KDU — rate editable in J
    r_kdu = r
    s(r, 3, "Kiti darbo užmokesčio priskaitymai")
    s(r,  9, f"=ROUND(ROUND(J{r1},2)*J{r_kdu},2)", bold=True, fmt=EUR_FMT)
    s(r, 10, OVERHEAD["KDU"], fmt="0.0%")
    s(r, 15, round(1 + OVERHEAD["PVM"], 4))  # O: VAT factor cell, as in the file
    s(r, 16, "KDU")
    r += 1

    # PMD — rate editable in K
    r_pmd = r
    s(r, 3, "Papildomų medžiagų vertė")
    s(r,  9, f"=ROUND(K{r_pmd}*ROUND(K{r1},2),2)", bold=True, fmt=EUR_FMT)
    s(r, 11, OVERHEAD["PMD"], fmt="0.0%")
    s(r, 16, "PMD")
    r += 1

    # PMZ — rate editable in L
    r_pmz = r
    s(r, 3, "Papildomų mechanizmų vertė")
    s(r,  9, f"=ROUND(L{r_pmz}*ROUND(L{r1},2),2)", bold=True, fmt=EUR_FMT)
    s(r, 12, OVERHEAD["PMZ"], fmt="0.0%")
    s(r, 16, "PMZ")
    r += 1

    # SOD — rate editable in J, applied on labor incl. KDU
    r_sod = r
    s(r, 3, "Soc. draudimas")
    s(r,  9, f"=ROUND(ROUND(J{r1},2)*J{r_sod}*(1+J{r_kdu}),2)", bold=True, fmt=EUR_FMT)
    s(r, 10, OVERHEAD["SOD"], fmt="0.0%")
    s(r, 16, "SOD")
    r += 1

    # #2
    r2 = r
    s(r, 2, "Iš viso #2 (išlaidos statinio statybos darbams)", bold=True)
    s(r,  9, f"=SUM(I{r1}:I{r_sod})", bold=True, fmt=EUR_FMT)
    s(r, 10, f"=ROUND(J{r_kdu}*ROUND(J{r1},2)+ROUND(J{r1},2)+I{r_sod},2)", fmt="0.00")
    s(r, 11, f"=ROUND(ROUND(K{r1},2)*K{r_pmd}+ROUND(K{r1},2),2)", fmt="0.00")
    s(r, 12, f"=ROUND(L{r_pmz}*ROUND(L{r1},2)+ROUND(L{r1},2),2)", fmt="0.00")
    r += 1

    # STI — rates editable in J/K/L
    r_sti = r
    s(r, 3, "Statybvietės išlaidos")
    s(r,  9, f"=ROUND(J{r_sti}*J{r2},2)+ROUND(K{r_sti}*K{r2},2)+ROUND(L{r_sti}*L{r2},2)",
      bold=True, fmt=EUR_FMT)
    s(r, 10, OVERHEAD["STI"], fmt="0.0%")
    s(r, 11, OVERHEAD["STI"], fmt="0.0%")
    s(r, 12, OVERHEAD["STI"], fmt="0.0%")
    s(r, 16, "STI")
    r += 1

    # #3
    r3 = r
    s(r, 2, "Iš viso #3 (tiesioginės išlaidos)", bold=True)
    s(r,  9, f"=J{r3}+K{r3}+L{r3}", bold=True, fmt=EUR_FMT)
    s(r, 10, f"=ROUND(J{r2}*J{r_sti}+J{r2},2)", fmt="0.00")
    s(r, 11, f"=ROUND(K{r2}*K{r_sti}+K{r2},2)", fmt="0.00")
    s(r, 12, f"=ROUND(L{r2}*L{r_sti}+L{r2},2)", fmt="0.00")
    r += 1

    # Indeksas — editable coefficients
    r_ind = r
    s(r, 3, "Indeksas")
    s(r, 10, 1)
    s(r, 11, 1)
    s(r, 12, 1)
    r += 1

    # Po indeksacijos
    r_ind_tot = r
    s(r, 2, "Po indeksacijos iš viso", bold=True)
    s(r,  9, f"=J{r_ind_tot}+K{r_ind_tot}+L{r_ind_tot}", bold=True, fmt=EUR_FMT)
    s(r, 10, f"=ROUND(J{r_ind}*J{r3},2)", fmt="0.00")
    s(r, 11, f"=ROUND(K{r_ind}*K{r3},2)", fmt="0.00")
    s(r, 12, f"=ROUND(L{r_ind}*L{r3},2)", fmt="0.00")
    r += 1

    # PRI — rate editable in J, applied on labor incl. KDU
    r_pri = r
    s(r, 3, "Pridėtinės išlaidos")
    s(r,  9, f"=ROUND((J{r1}*J{r_kdu}+J{r1})*J{r_pri},2)", bold=True, fmt=EUR_FMT)
    s(r, 10, OVERHEAD["PRI"], fmt="0.0%")
    s(r, 16, "PRI")
    s(r, 18, OVERHEAD["PRI"], h="center")
    r += 1

    # PLN — rates editable in J/K/L
    r_pln = r
    s(r, 3, "Pelnas")
    s(r,  9, f"=ROUND(J{r_pln}*(ROUND(J{r_pri}*(J{r1}+J{r_kdu}*J{r1}),2)+J{r_ind_tot}),2)"
             f"+ROUND(K{r_pln}*(ROUND(K{r_pri}*K{r2}+K{r_ind_tot},2)),2)"
             f"+ROUND(L{r_pln}*(ROUND(L{r_pri}*L{r2}+L{r_ind_tot},2)),2)",
      bold=True, fmt=EUR_FMT)
    s(r, 10, OVERHEAD["PLN"], fmt="0.0%")
    s(r, 11, OVERHEAD["PLN"], fmt="0.0%")
    s(r, 12, OVERHEAD["PLN"], fmt="0.0%")
    s(r, 16, "PLN")
    r += 1

    # #4
    r4 = r
    s(r, 2, "Iš viso #4 (su netiesioginėmis išlaidomis)", bold=True)
    s(r,  9, f"=J{r4}+K{r4}+L{r4}", bold=True, fmt=EUR_FMT)
    s(r, 10, f"=ROUND(J{r_pln}*(ROUND(J{r_pri}*(J{r1}+J{r_kdu}*J{r1}),2)+J{r_ind_tot})"
             f"+(ROUND(J{r_pri}*(J{r1}+J{r_kdu}*J{r1}),2)+J{r_ind_tot}),2)", fmt="0.00")
    s(r, 11, f"=ROUND(K{r_pln}*(ROUND(K{r_pri}*K{r2}+K{r_ind_tot},2))"
             f"+(ROUND(K{r_pri}*K{r2}+K{r_ind_tot},2)),2)", fmt="0.00")
    s(r, 12, f"=ROUND(L{r_pln}*(ROUND(L{r_pri}*L{r2}+L{r_ind_tot},2))"
             f"+(ROUND(L{r_pri}*L{r2}+L{r_ind_tot},2)),2)", fmt="0.00")
    r += 1

    # PVM — rates editable in J/K/L
    r_pvm = r
    s(r, 3, "PVM")
    s(r,  9, f"=ROUND(J{r4}*J{r_pvm}+K{r4}*K{r_pvm}+L{r4}*L{r_pvm},2)", bold=True, fmt=EUR_FMT)
    s(r, 10, OVERHEAD["PVM"], fmt="0.0%")
    s(r, 11, OVERHEAD["PVM"], fmt="0.0%")
    s(r, 12, OVERHEAD["PVM"], fmt="0.0%")
    s(r, 16, "PVM")
    r += 1

    # #5
    r5 = r
    assert r5 == _r5, f"Row-5 mismatch: expected {_r5}, got {r5}"
    s(r, 2, "Iš viso #5 (kaina su PVM)", bold=True)
    s(r,  9, f"=I{r4}+I{r_pvm}", bold=True, fmt=EUR_FMT)
    s(r, 10, f"=IF(J{r4}<>0,I{r5}-K{r5}-L{r5},0)", fmt="0.00")
    s(r, 11, f"=ROUND(K{r_pvm}*K{r4}+K{r4},2)", fmt="0.00")
    s(r, 12, f"=ROUND(L{r_pvm}*L{r4}+L{r4},2)", fmt="0.00")
    s(r, 14, "Pabaiga")
    r += 2

    s(r, 3, "Sudarė:")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── HTML preview ─────────────────────────────────────────────────────────────
def render_samata_html(matched_items, project_name=""):
    t = compute_totals(matched_items)
    from datetime import date

    # Column definitions: (label, width, align) — matches Excel column layout A–M
    cols = [
        ("Nr.",               "40px",  "center"),
        ("Darbo pavadinimas", "300px", "left"),
        ("Kodas",             "90px",  "left"),
        ("Mat. vnt",          "70px",  "center"),
        ("Norma",             "60px",  "right"),
        ("Koef",              "15px",  "right"),
        ("Kaina",             "80px",  "right"),
        ("Kiekis",            "70px",  "right"),
        ("Suma",              "110px", "right"),
        ("Darbas",            "100px", "right"),
        ("Medžiagos",         "100px", "right"),
        ("Mechanizmai",       "100px", "right"),
        ("Subrangovai",       "80px",  "right"),
    ]

    BASE = "font-family:'Times New Roman',serif; font-size:9.75pt;"
    HDR  = f"{BASE} font-weight:bold; background:#fff; border-bottom:2px solid #000; text-align:center; padding:4px 6px; white-space:nowrap;"
    TD   = f"{BASE} padding:2px 6px; border-bottom:1px solid #ddd; vertical-align:top;"
    BOLD = f"{BASE} font-weight:bold; padding:2px 6px; border-bottom:1px solid #bbb; vertical-align:top;"
    TOT  = f"{BASE} font-weight:bold; padding:3px 6px; border-top:2px solid #000; border-bottom:2px solid #000; vertical-align:top; background:#f5f5f5;"
    SEC  = f"{BASE} font-weight:bold; padding:4px 6px; border-bottom:1px solid #999; background:#fff; vertical-align:top;"
    CHG  = f"{BASE} padding:2px 6px; border-bottom:1px solid #eee; vertical-align:top; color:#444;"

    def th(label, width, align):
        return f'<th style="{HDR} width:{width}; text-align:{align};">{label}</th>'

    def td(val, style, align="left", colspan=1):
        sp = f' colspan="{colspan}"' if colspan > 1 else ""
        return f'<td style="{style} text-align:{align};"{sp}>{val if val is not None else ""}</td>'

    rows_html = []

    # ── Header
    rows_html.append(f'''
    <tr>
      <td colspan="13" style="font-family:'Times New Roman',serif; font-size:11pt;
          font-weight:bold; padding:8px 6px; border-bottom:1px solid #ccc;">
        {project_name or "—"}
        <span style="float:right; font-size:9pt; font-weight:normal;">
          L o k a l i n ė &nbsp; s ą m a t a &nbsp;&nbsp;|&nbsp;&nbsp;
          Sudaryta {date.today().strftime("%Y.%m")} kainų lygiu &nbsp;&nbsp;|&nbsp;&nbsp;
          Iš viso su PVM: <strong>{lt_num(t["t5"])} €</strong>
        </span>
      </td>
    </tr>''')

    # ── Column headers
    rows_html.append('<tr>' + ''.join(th(l, w, a) for l, w, a in cols) + '</tr>')

    # ── Section header
    section_name = "Darbai"
    rows_html.append(f'''
    <tr>
      <td style="{SEC}"></td>
      <td style="{SEC}" colspan="12">Skyrius &nbsp; {section_name}</td>
    </tr>''')

    NA_ROW = f"{BASE} font-weight:bold; padding:2px 6px; border-bottom:1px solid #cc0000; vertical-align:top; background:#FF9999; color:#000;"

    # ── Work items
    item_no = 1
    for item in matched_items:
        no_match   = item.get("status") == "no_match" or not item.get("matched_resources")
        kiekis     = item.get("kiekis") or 0
        item_total = item.get("total_cost") or 0
        item_labor = item.get("total_labor") or 0
        item_mat   = item.get("total_materials") or 0
        item_mach  = item.get("total_machinery") or 0
        unit_price = round(item_total / kiekis, 2) if kiekis and not no_match else 0

        row_style = NA_ROW if no_match else BOLD
        na = "NA"

        rows_html.append(
            '<tr>'
            + td(item_no,                                row_style, "center")
            + td(item.get("pavadinimas",""),             row_style, "left")
            + td(item.get("eil_nr",""),                  row_style, "left")
            + td(item.get("vnt",""),                     row_style, "center")
            + td("",                                     row_style, "right")  # Norma
            + td("",                                     row_style, "right")  # Koef
            + td("" if no_match else lt_num(unit_price), row_style, "right")  # Kaina
            + td(kiekis,                                 row_style, "right")  # Kiekis
            + td(na if no_match else lt_num(item_total), row_style, "right")  # Suma
            + td(na if no_match else lt_num(item_labor), row_style, "right")  # Darbas
            + td(na if no_match else lt_num(item_mat),   row_style, "right")  # Medžiagos
            + td(na if no_match else lt_num(item_mach),  row_style, "right")  # Mechanizmai
            + td("",                                     row_style, "right")  # Subrangovai
            + '</tr>'
        )

        for res in item.get("matched_resources", []):
            cost = res.get("total_cost") or 0
            cat  = res.get("category", "")
            rows_html.append(
                '<tr>'
                + td("",                              TD,   "center")
                + td(res.get("name",""),              TD,   "left")
                + td(res.get("code",""),              TD,   "left")
                + td(res.get("unit",""),              TD,   "center")
                + td(res.get("qty_per_spec_unit",""), TD,   "right")  # Norma
                + td("",                              TD,   "right")  # Koef
                + td(res.get("unit_price",""),        TD,   "right")  # Kaina
                + td(round(res.get("total_qty") or 0, 3), TD, "right")  # Kiekis
                + td(lt_num(cost),                    TD,   "right")  # Suma
                + td(lt_num(cost) if cat=="labor"     else "", TD, "right")  # Darbas
                + td(lt_num(cost) if cat=="material"  else "", TD, "right")  # Medžiagos
                + td(lt_num(cost) if cat=="machinery" else "", TD, "right")  # Mechanizmai
                + td("",                              TD,   "right")  # Subrangovai
                + '</tr>'
            )
        item_no += 1

    # ── Section total
    rows_html.append(
        '<tr>'
        + td("",                      TOT, "left", colspan=2)
        + td("Iš viso už skyrių",    TOT, "left", colspan=6)
        + td(lt_num(t["direct"]),    TOT, "right")
        + td(lt_num(t["labor"]),     TOT, "right")
        + td(lt_num(t["materials"]), TOT, "right")
        + td(lt_num(t["machinery"]), TOT, "right")
        + td("",                     TOT, "right")
        + '</tr><tr><td colspan="13" style="height:10px;"></td></tr>'
    )

    # ── Grand totals helper
    def tot_row(label, i, j="", k="", l="", bold=True):
        sty = TOT if bold else CHG
        return (
            '<tr>'
            + td("",    sty, colspan=2)
            + td(label, sty, "left", colspan=6)
            + td(i,     sty, "right")
            + td(j,     sty, "right")
            + td(k,     sty, "right")
            + td(l,     sty, "right")
            + td("",    sty, "right")
            + '</tr>'
        )

    rows_html += [
        tot_row("Iš viso #1",
                lt_num(t["direct"]), lt_num(t["labor"]),
                lt_num(t["materials"]), lt_num(t["machinery"])),
        tot_row(f"Kiti darbo užmokesčio priskaitymai (KDU {OVERHEAD['KDU']*100:.0f}%)",
                lt_num(t["KDU"]), lt_num(t["KDU"]), bold=False),
        tot_row(f"Papildomų medžiagų vertė (PMD {OVERHEAD['PMD']*100:.0f}%)",
                lt_num(t["PMD"]), k=lt_num(t["PMD"]), bold=False),
        tot_row(f"Papildomų mechanizmų vertė (PMZ {OVERHEAD['PMZ']*100:.0f}%)",
                lt_num(t["PMZ"]), l=lt_num(t["PMZ"]), bold=False),
        tot_row(f"Soc. draudimas (SOD {OVERHEAD['SOD']*100:.2f}%)",
                lt_num(t["SOD"]), lt_num(t["SOD"]), bold=False),
        tot_row("Iš viso #2 (išlaidos statinio statybos darbams)",
                lt_num(t["t2"]),
                lt_num(t["labor"]+t["KDU"]+t["SOD"]),
                lt_num(t["materials"]+t["PMD"]),
                lt_num(t["machinery"]+t["PMZ"])),
        tot_row(f"Statybvietės išlaidos (STI {OVERHEAD['STI']*100:.0f}%)",
                lt_num(t["STI"]), bold=False),
        tot_row("Iš viso #3 (tiesioginės išlaidos)", lt_num(t["t3"])),
        tot_row(f"Pridėtinės išlaidos (PRI {OVERHEAD['PRI']*100:.1f}%)",
                lt_num(t["PRI"]), lt_num(t["PRI"]), bold=False),
        tot_row(f"Pelnas (PLN {OVERHEAD['PLN']*100:.0f}%)",
                lt_num(t["PLN"]), bold=False),
        tot_row("Iš viso #4 (su netiesioginėmis išlaidomis)", lt_num(t["t4"])),
        tot_row(f"PVM {OVERHEAD['PVM']*100:.0f}%", lt_num(t["PVM"]), bold=False),
    ]

    # Final total — highlighted
    rows_html.append(f'''
    <tr style="background:#1a3a6b;">
      <td colspan="2" style="{BASE} color:#fff; font-weight:bold; padding:5px 6px;"></td>
      <td colspan="6" style="{BASE} color:#fff; font-weight:bold; padding:5px 6px;">
        Iš viso #5 (kaina su PVM)
      </td>
      <td style="{BASE} color:#fff; font-weight:bold; padding:5px 6px; text-align:right;">
        {lt_num(t["t5"])}
      </td>
      <td colspan="4" style="{BASE} color:#fff; padding:5px 6px;"></td>
    </tr>''')

    col_group = "".join(
        f'<col style="width:{w}">' for _, w, _ in cols
    )

    return f"""
    <div style="overflow-x:auto; font-family:'Times New Roman',serif;">
      <table style="border-collapse:collapse; width:100%; min-width:900px;">
        <colgroup>{col_group}</colgroup>
        <tbody>{''.join(rows_html)}</tbody>
      </table>
    </div>
    """


# ── Streamlit app ─────────────────────────────────────────────────────────────
import streamlit.components.v1 as components

st.set_page_config(page_title="Sąmatos skaičiavimas", layout="wide", page_icon="🏗️")

st.markdown("""
<style>
  .block-container {
      max-width: 1100px !important;
      padding-left: 2rem !important;
      padding-right: 2rem !important;
  }
</style>
""", unsafe_allow_html=True)

resources_df = load_resources()

st.title("Sąmatos skaičiavimas")
st.caption(f"Kainų duomenų bazė: {len(resources_df)} resursai")

# ── Inputs ────────────────────────────────────────────────────────────────────
st.divider()

uploaded_file = st.file_uploader("PDF projekto specifikacija", type=["pdf"])
col1, col2 = st.columns(2)
with col1:
    start_page = st.number_input("Puslapis nuo", min_value=1, value=1)
with col2:
    end_page = st.number_input("Puslapis iki", min_value=1, value=1)

project_name = st.text_input("Projekto pavadinimas",
                              placeholder="pvz. Sandėlio pastatas, Vilnius, 2024")

if st.button("Generuoti sąmatą", type="primary", use_container_width=True):
    if not uploaded_file:
        st.error("Įkelkite PDF failą.")
    elif start_page > end_page:
        st.error("Puslapis 'nuo' negali būti didesnis už 'iki'.")
    else:
        pdf_bytes = uploaded_file.read()
        total_pages = get_pdf_page_count(pdf_bytes)
        if start_page > total_pages:
            st.error(f"Šiame PDF yra tik {total_pages} puslapiai (-ų). "
                     f"Pasirinkote puslapį {start_page}. Pataisykite puslapių intervalą.")
            st.stop()
        if end_page > total_pages:
            st.warning(f"PDF turi {total_pages} puslapius (-ų) — intervalo pabaiga "
                       f"sumažinta nuo {end_page} iki {total_pages}.")
            end_page = total_pages

        n_pages = end_page - start_page + 1
        with st.spinner(f"1/3 · Ištraukiamos lentelės iš {n_pages} puslapių lygiagrečiai..."):
            images = extract_specific_pages_as_images(pdf_bytes, start_page, end_page)

            def _extract_page(img):
                result = extract_table_data(img["base64_image"])
                return img["page_num"], result.get("table_data", [])

            page_results = {}
            errors = []
            if not images:
                st.error("Nepavyko nuskaityti nurodytų puslapių iš PDF.")
                st.stop()
            with ThreadPoolExecutor(max_workers=min(len(images), 10)) as pool:
                futures = {pool.submit(_extract_page, img): img["page_num"] for img in images}
                for future in as_completed(futures):
                    page_num = futures[future]
                    try:
                        pnum, data = future.result()
                        page_results[pnum] = data
                    except Exception as e:
                        errors.append(f"Puslapis {page_num}: {e}")

            for err in errors:
                st.error(f"Klaida — {err}")

            # Reassemble pages in order
            all_data = []
            for pnum in sorted(page_results):
                all_data.extend(page_results[pnum])

        if not all_data:
            st.error("Nepavyko ištraukti duomenų iš PDF.")
        else:
            st.session_state["parsed_data"] = all_data
            st.session_state["project_name"] = project_name

            # Estimator planning: normative naming, unit conversion (kg→t, m2→100 m2),
            # splitting kompleksiniai works, adding implied works (šiukšlių išvežimas etc.)
            with st.spinner(f"2/3 · Sudaromas sąmatos darbų sąrašas ({len(all_data)} žiniaraščio eilučių)..."):
                try:
                    planned = plan_work_items(all_data)
                    planned = refine_planned_items(planned, all_data)
                    # Feed the matcher estimator-style items; the suggested normative
                    # code goes to "eil_nr", which the Excel writes into column C.
                    work_items = [{
                        "eil_nr":      p.get("kodas", ""),
                        "zin_nr":      p.get("zin_nr", ""),
                        "pavadinimas": p.get("pavadinimas", ""),
                        "vnt":         p.get("vnt", ""),
                        "kiekis":      p.get("kiekis"),
                    } for p in planned]
                    st.session_state["planned"] = work_items
                except Exception as e:
                    st.warning(f"Darbų sąrašo sudarymas nepavyko ({e}) — naudojamos žiniaraščio eilutės tiesiogiai.")
                    work_items = all_data

            n_items = len(work_items)
            n_batches = max(1, (n_items + MATCH_BATCH_SIZE - 1) // MATCH_BATCH_SIZE)
            with st.spinner(f"3/3 · AI derina {n_items} pozicijų ({n_batches} lygiagrečių partijų)..."):
                try:
                    parsed_df = pd.DataFrame(work_items)
                    matched = match_items_to_resources(parsed_df, resources_df)
                    st.session_state["matched"] = matched
                except Exception as e:
                    st.error(f"Klaida derinant: {e}")
                    st.stop()

if "matched" in st.session_state:
    matched = st.session_state["matched"]
    t = compute_totals(matched)
    na_count = sum(1 for i in matched if i.get("status") == "no_match" or not i.get("matched_resources"))

    if na_count:
        st.warning(f"**{na_count}** pozicija (-os) nepavyko suderinti su duomenų baze ir pažymėtos **NA**.")

# ── Preview & download ────────────────────────────────────────────────────────
if "matched" in st.session_state:
    matched      = st.session_state["matched"]
    project_name = st.session_state.get("project_name", "")

    st.divider()
    st.subheader("Sąmatos peržiūra ir atsisiuntimas")

    na_count = sum(1 for i in matched if i.get("status") == "no_match" or not i.get("matched_resources"))
    if na_count:
        st.info(f"{na_count} eilutė (-ės) pažymėta NA — nerasta atitikmenų kainų duomenų bazėje.")

    html_preview = render_samata_html(matched, project_name)
    components.html(html_preview, height=800, scrolling=True)

    excel_buf = generate_samata_excel(matched, project_name=project_name)
    st.download_button(
        "Atsisiųsti sąmatą (.xlsx)",
        data=excel_buf,
        file_name="samata.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )
